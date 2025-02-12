# @name         Arona AI Helper
# @version      v0.1
# @description  Geather student usage in different raids
# @author       Jacky Ho (javascript code)
# @author       fiseleo (python script)


import requests
import time
import re
from openpyxl import Workbook


def get_json(url):
    """用 requests 抓取 JSON 資料，失敗時傳回 None"""
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Failed to fetch {url} (status code: {response.status_code})")
            return None
    except Exception as e:
        print(f"Exception while fetching {url}: {e}")
        return None


def sort_students(students):
    """
    將學生資料排序，優先依 isLimited（True 在前）再依 id（數值較大在前）
    備註：假設學生資料中 id 為數字或可轉換成整數
    """
    def sort_key(item):
        return (1 if item.get("isLimited") else 0, int(item.get("id", 0)))
    return sorted(students, key=sort_key, reverse=True)


def is_same_raid(a, b):
    """
    判斷兩場 raid 是否相同：
      - RaidId 必須相同
      - Terrain 必須相同
      - 若存在 ArmorTypes，則 a 中的每個類型都必須出現在 b 的 ArmorTypes 中
    """
    if a.get("RaidId") != b.get("RaidId"):
        return False
    if a.get("Terrain") != b.get("Terrain"):
        return False
    if "ArmorTypes" in a:
        for t in a.get("ArmorTypes", []):
            if t not in b.get("ArmorTypes", []):
                return False
    return True




def format_sheet_name(s):
    """
    將工作表名稱中的非法字元移除，
    並限制名稱長度不超過 31 個字元（Excel 限制）
    """
    s = re.sub(r'[\/\\\?\*\[\]＊]', '', s)
    s = s.strip()
    return s[:31]


def main():
    # 設定各項 URL 與參數（參考原本的 factInfo）
    fact_info = {
        "eraid": {"id": 10, "endDate": "2024-05-22 03:59", "min": 1},
        "raid": {"id": 61, "endDate": "2024-01-10 03:59", "min": 47},
        "minRaidSeparateDay": 28,
        "lagDay": 182,
        "summaryRank": ["1000", "5000", "10000", "20000"],
        "raidUrl": "https://media.arona.ai/data/v3/raid/<id>/total",
        "eraidUrl": "https://media.arona.ai/data/v3/eraid/<id>/total",
        "raidInfo": "https://schaledb.com/data/tw/raids.json",
        "studentUrl": "https://schaledb.com/data/tw/students.json"
    }

    # 取得學生資料
    std_info = get_json(fact_info["studentUrl"])
    if std_info is None:
        print("Fail to fetch student info!")
        return

    # 取得 raid 資料
    raid_info = get_json(fact_info["raidInfo"])
    if raid_info is None:
        print("Fail to fetch raid info!")
        return

    # 處理目前 TW 服的 Raid 賽季
    try:
        curr_tw_raid = raid_info["RaidSeasons"][1]["Seasons"][-1]
        if curr_tw_raid["End"] > time.time():
            curr_tw_raid = raid_info["RaidSeasons"][1]["Seasons"][-2]
    except Exception as e:
        print("Error processing current Taiwan raid season:", e)
        return

    jp_raids = raid_info["RaidSeasons"][0]["Seasons"]
    raid_map = {}
    ref_jp_raid = len(jp_raids) - 1
    while ref_jp_raid >= 0 and not is_same_raid(jp_raids[ref_jp_raid], curr_tw_raid):
        curr_raid = jp_raids[ref_jp_raid]
        try:
            raid_name = raid_info["Raid"][curr_raid["RaidId"] - 1]["Name"] + " " + (curr_raid["Terrain"])
        except Exception as e:
            print("Error processing raid name:", e)
            raid_name = ""
        # 以 SeasonDisplay 作為 key，值儲存名稱資訊
        raid_map[curr_raid["SeasonDisplay"]] = {"name": raid_name}
        ref_jp_raid -= 1

    # 處理目前 TW 服的 ERAID 賽季
    try:
        curr_tw_eraid = raid_info["RaidSeasons"][1]["EliminateSeasons"][-1]
        if curr_tw_eraid["End"] > time.time():
            curr_tw_eraid = raid_info["RaidSeasons"][1]["EliminateSeasons"][-2]
    except Exception as e:
        print("Error processing current Taiwan eraid season:", e)
        return

    jp_eraids = raid_info["RaidSeasons"][0]["EliminateSeasons"]
    eraid_map = {}
    ref_jp_eraid = len(jp_eraids) - 1
    while ref_jp_eraid >= 0 and not is_same_raid(jp_eraids[ref_jp_eraid], curr_tw_eraid):
        curr_eraid = jp_eraids[ref_jp_eraid]
        try:
            eraid_name = raid_info["Raid"][curr_eraid["RaidId"] - 1]["Name"] + " " + (curr_eraid["Terrain"])
        except Exception as e:
            print("Error processing eraid name:", e)
            eraid_name = ""
        eraid_map[curr_eraid["SeasonDisplay"]] = {"name": eraid_name}
        ref_jp_eraid -= 1

    # 分別依 raid 與 eraid 取得資料
    search_raid = {}
    for raid_id in raid_map.keys():
        url = fact_info["raidUrl"].replace("<id>", str(raid_id))
        print("Getting raid info for", raid_id, raid_map[raid_id]["name"])
        retrieved_info = get_json(url)
        if retrieved_info is not None:
            search_raid[raid_id] = retrieved_info

    search_eraid = {}
    for eraid_id in eraid_map.keys():
        url = fact_info["eraidUrl"].replace("<id>", str(eraid_id))
        print("Getting eraid info for", eraid_id, eraid_map[eraid_id]["name"])
        retrieved_info = get_json(url)
        if retrieved_info is not None:
            search_eraid[eraid_id] = retrieved_info

    student_map = {}  # 用來儲存每位學生在各場戰役的詳細資料
    rank_map = {}     # 用來彙整每個階層的資料
    time_map = {}     # 用來儲存各場戰役對應的時間標記

    # 處理 raid 資料
    for no_raid, curr_raid_info in search_raid.items():
        raid_name = f"S{no_raid} - {raid_map[no_raid]['name']} 總力戰"
        try:
            key_time = str(curr_raid_info["trophyCutByTime"]["id"][0])
        except Exception:
            key_time = str(no_raid)
        time_map[key_time] = raid_name

        char_usage = curr_raid_info.get("characterUsage", {}).get("r", {})
        for rank_range, std_dict in char_usage.items():
            for std_id, usage_list in std_dict.items():
                std_entry = std_info.get(std_id, {})
                std_nm = std_entry.get("Name", "")
                is_limited = std_entry.get("IsLimited", False)
                if rank_range in fact_info["summaryRank"]:
                    rank_index = fact_info["summaryRank"].index(rank_range)
                    if rank_range not in rank_map:
                        rank_map[rank_range] = {}
                    if std_id not in rank_map[rank_range]:
                        rank_map[rank_range][std_id] = {"id": std_id, "stdNm": std_nm, "max": -1, "isLimited": is_limited, "cnt": 0}
                    use_cnt = sum(usage_list)
                    if rank_index != 0:
                        prev_rank = fact_info["summaryRank"][rank_index - 1]
                        if prev_rank in char_usage and std_id in char_usage[prev_rank]:
                            use_cnt -= sum(char_usage[prev_rank][std_id])
                    if use_cnt > 0:
                        rank_map[rank_range][std_id][raid_name] = use_cnt
                        rank_map[rank_range][std_id]["max"] = max(rank_map[rank_range][std_id]["max"], use_cnt)
                        rank_map[rank_range][std_id]["cnt"] += 1

                if std_id not in student_map:
                    student_map[std_id] = {}
                if raid_name not in student_map[std_id]:
                    student_map[std_id][raid_name] = {}
                student_map[std_id][raid_name][rank_range] = usage_list

    # 處理 eraid 資料
    for no_eraid, curr_eraid_info in search_eraid.items():
        try:
            eraid_time = curr_eraid_info["trophyCutByTime"]["id"][0]
        except Exception:
            eraid_time = str(no_eraid)
        char_usage_all = curr_eraid_info.get("characterUsage", {})
        for battle_type, battle_data in char_usage_all.items():
            last_under_index = battle_type.rfind("_")
            battle_suffix = battle_type[last_under_index+1:]
            eraid_name = f"S{no_eraid} - {eraid_map[no_eraid]['name']} {battle_suffix} 大決戰"
            # 為了避免 key 重複，將 eraid_time 與 eraid_name 連接起來作為 key
            time_map[str(eraid_time) + eraid_name] = eraid_name
            char_usage = battle_data.get("r", {})
            for rank_range, std_dict in char_usage.items():
                for std_id, usage_list in std_dict.items():
                    std_entry = std_info.get(std_id, {})
                    std_nm = std_entry.get("Name", "")
                    is_limited = std_entry.get("IsLimited", False)
                    if rank_range in fact_info["summaryRank"]:
                        rank_index = fact_info["summaryRank"].index(rank_range)
                        if rank_range not in rank_map:
                            rank_map[rank_range] = {}
                        if std_id not in rank_map[rank_range]:
                            rank_map[rank_range][std_id] = {"id": std_id, "stdNm": std_nm, "max": -1, "isLimited": is_limited, "cnt": 0}
                        use_cnt = sum(usage_list)
                        if rank_index != 0:
                            prev_rank = fact_info["summaryRank"][rank_index - 1]
                            if prev_rank in char_usage and std_id in char_usage[prev_rank]:
                                use_cnt -= sum(char_usage[prev_rank][std_id])
                        if use_cnt > 0:
                            rank_map[rank_range][std_id][eraid_name] = use_cnt
                            rank_map[rank_range][std_id]["max"] = max(rank_map[rank_range][std_id]["max"], use_cnt)
                            rank_map[rank_range][std_id]["cnt"] += 1
                    if std_id not in student_map:
                        student_map[std_id] = {}
                    if eraid_name not in student_map[std_id]:
                        student_map[std_id][eraid_name] = {}
                    student_map[std_id][eraid_name][rank_range] = usage_list

    # 建立 Excel 工作簿
    wb = Workbook()
    # 刪除預設的工作表
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_array = ["id", "stdNm", "isLimited", "cnt", "max"]
    raid_array_by_date = []
    # 依照 time_map key 的排序順序加入欄位（順序與原 JS 程式類似）
    for t in sorted(time_map.keys()):
        header_array.append(time_map[t])
        raid_array_by_date.append(time_map[t])

    # 建立依各階層統計的 Summary 工作表
    # 依照 fact_info["summaryRank"] 的順序處理（若該階層有資料）
    previous_rank = None
    for i, rank_range in enumerate(fact_info["summaryRank"]):
        if rank_range not in rank_map:
            continue
        data_array = list(rank_map[rank_range].values())
        data_array = sort_students(data_array)
        if i == 0:
            sheet_name = f"Summary - Rank {rank_range}"
        else:
            sheet_name = f"Summary - Rank {previous_rank} to {rank_range}"
        previous_rank = rank_range
        ws = wb.create_sheet(title=format_sheet_name(sheet_name))
        # 寫入表頭
        ws.append(header_array)
        # 寫入資料列（依 header_array 的順序取值）
        for row_data in data_array:
            row = [row_data.get(col, "") for col in header_array]
            ws.append(row)

    # 建立每位學生的詳細資料工作表
    # 每個工作表中依 raid_array_by_date 的順序顯示各場戰役的明細
    for std_id, raids in student_map.items():
        data_array = []
        for raid_name in raid_array_by_date:
            if raid_name in raids:
                # 第一列顯示戰役名稱
                data_array.append({"1": raid_name})
                # 第二列顯示該戰役的表頭
                data_array.append({
                    "1": "排名",
                    "2": "借用",
                    "3": "三星以下",
                    "4": "四星",
                    "5": "五星無武",
                    "6": "專一",
                    "7": "專二",
                    "8": "專三",
                    "9": "共計"
                })
                # 依照各階層顯示數值，若無資料則填 0
                for rank in fact_info["summaryRank"]:
                    if rank in raids[raid_name]:
                        use_arr = raids[raid_name][rank]
                        total = sum(use_arr)
                        data_array.append({
                            "1": rank + "以下",
                            "2": use_arr[0] if len(use_arr) > 0 else 0,
                            "3": use_arr[1] if len(use_arr) > 1 else 0,
                            "4": use_arr[2] if len(use_arr) > 2 else 0,
                            "5": use_arr[3] if len(use_arr) > 3 else 0,
                            "6": use_arr[4] if len(use_arr) > 4 else 0,
                            "7": use_arr[5] if len(use_arr) > 5 else 0,
                            "8": use_arr[6] if len(use_arr) > 6 else 0,
                            "9": total
                        })
                    else:
                        data_array.append({
                            "1": rank + "以下",
                            "2": 0,
                            "3": 0,
                            "4": 0,
                            "5": 0,
                            "6": 0,
                            "7": 0,
                            "8": 0,
                            "9": 0
                        })
                # 空一列作區隔
                data_array.append({})
        # 這裡設定學生工作表的欄位順序
        student_header = ["1", "2", "3", "4", "5", "6", "7", "8", "9"]
        sheet_title = format_sheet_name(f"{std_id}")
        ws = wb.create_sheet(title=sheet_title)
        # 依序寫入每一列（不額外產生表頭）
        for row_data in data_array:
            row = [row_data.get(col, "") for col in student_header]
            ws.append(row)

    # 寫出 Excel 檔案
    output_filename = "data.xlsx"
    wb.save(output_filename)
    print(f"Excel file '{output_filename}' has been created.")


if __name__ == '__main__':
    main()